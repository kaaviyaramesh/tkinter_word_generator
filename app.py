import tkinter
from tkinter import OptionMenu, StringVar, Frame
import random
import openpyxl

frame = None

window = tkinter.Tk()
window.geometry('700x200')
window.title("words")


def get_words():
    global frame
    # print(label)
    if frame != None:
        # print("Hello")
        frame.destroy()
    value = clicked.get()
    sheet_to_use = wb[value]
    max_ = sheet_to_use.max_row
    # print(max_)
    all_words = []
    for i in range(1, max_+1):
        if sheet_to_use.cell(row=i, column=2).value == 0:
            all_words.append(sheet_to_use.cell(row=i, column=1).value)
    # print(all_words)

    iteration = 5
    if len(all_words) < iteration:
        iteration = len(all_words)
    # print("iteration",iteration)
    wordlist = []
    rand_num = random.sample(range(0, len(all_words)), iteration)
    # print("rand",rand_num)
    frame = Frame(window)
    for i in rand_num:
        label = tkinter.Label(frame, text=all_words[i])
        label.pack()
    frame.pack()
    return


wb = openpyxl.load_workbook("test_list.xlsx")
sheet = wb.active
values = wb.sheetnames

clicked = StringVar()
clicked.set("Select")

drop = OptionMenu(window, clicked, *values)
drop.pack()


button = tkinter.Button(window, text="Go", command=get_words).pack()
window.mainloop()
